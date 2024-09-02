import os
import argparse
from bs4 import BeautifulSoup
import json
import mechanize
import pathlib
import re
import sys 
import time 
import xml.etree.ElementTree as ET
import openpyxl 
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Dictionaries to map arguments to values
schlagwortOptionen = {"all": 1, "min": 2, "exact": 3}


class HandelsRegister:
    def __init__(self, args):
        self.args = args
        self.browser = mechanize.Browser()

        self.browser.set_debug_http(args.debug)
        self.browser.set_debug_responses(args.debug)
        # self.browser.set_debug_redirects(True)

        self.browser.set_handle_robots(False)
        self.browser.set_handle_equiv(True)
        self.browser.set_handle_gzip(True)
        self.browser.set_handle_refresh(False)
        self.browser.set_handle_redirect(True)
        self.browser.set_handle_referer(True)

        self.browser.addheaders = [
            (
                "User-Agent",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.5 Safari/605.1.15",
            ),
            ("Accept-Language", "en-GB,en;q=0.9"),
            ("Accept-Encoding", "gzip, deflate, br"),
            (
                "Accept",
                "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            ),
            ("Connection", "keep-alive"),
        ]

        self.cachedir = pathlib.Path("cache")
        self.cachedir.mkdir(parents=True, exist_ok=True) 

        self.externalexceldir = pathlib.Path("external_excel")
        self.externalexceldir.mkdir(parents=True, exist_ok=True) 

    def open_startpage(self):
        self.browser.open(
            "https://www.handelsregister.de/rp_web/welcome.xhtml", timeout=10
        )

    def companyname2cachename(self, companyname):
        # map a companyname to a filename, that caches the downloaded HTML, so re-running this script touches the
        # webserver less often.
        return self.cachedir / companyname 
    
    def externalexcelname(self, filename):
        return self.externalexceldir / filename

    def search_company_ies(self): 
        # Catch the situation if there's more than one company
        # Check for external excel file with company names
        externalexcelfilename = self.externalexcelname(self.args.externalExcel) 

        if self.args.loadExternalExcel == "true" and externalexcelfilename.exists(): 

            print("Please note that depending on the number of companies you have provided on the Excel sheet, it would take approximately 30 seconds to retrieve and process their detailed director(s) information for each company. Please close all Excel files while running this program.") 

            if os.path.exists(externalexcelfilename):
                # Traverse (read) the Excel file to retrieve the company names and compute on them
                companies_list_in_excel = openpyxl.load_workbook(externalexcelfilename) 

                worksheet = companies_list_in_excel.active

                print("Sheet names:", companies_list_in_excel.sheetnames)
                print("Sheet:", worksheet) 

                # cell_range = worksheet['A1'] 

                # print(cell_range) 

                for row in worksheet.iter_rows(min_row=1, max_col=1):
                    for cell in row:
                        # print(cell)
                        print(cell.value) 

                        # Check "cache" folder for the company name, in case it has already been searched for
                        cachename = self.companyname2cachename(cell.value) 

                        if self.args.force == False and cachename.exists():
                            with open(cachename, "r") as f:
                                html = f.read()
                                print("return cached content for %s" % cell.value) 
                        else:
                            self.browser.open(
                                "https://www.handelsregister.de/rp_web/erweitertesuche.xhtml"
                            )
                            if self.args.debug == True:
                                print(self.browser.title())

                            self.browser.select_form(name="form")

                            self.browser["form:schlagwoerter"] = cell.value
                            so_id = schlagwortOptionen.get(self.args.schlagwortOptionen)

                            self.browser["form:schlagwortOptionen"] = [str(so_id)]

                            response_result = self.browser.submit()

                            if self.args.debug == True:
                                print(self.browser.title())

                            html = response_result.read().decode("utf-8")
                            with open(cachename, "w") as f:
                                f.write(html) 

                        get_companies_in_searchresults(html) 


            else:
                print("File not found. Check that you got the file name correctly and that it is placed in the \"external_excel\" folder.") 



        elif self.args.loadExternalExcel == "false": 
            # Check "cache" folder for the company name, in case it has already been searched for
            cachename = self.companyname2cachename(self.args.schlagwoerter) 

            if self.args.force == False and cachename.exists():
                with open(cachename, "r") as f:
                    html = f.read()
                    print("return cached content for %s" % self.args.schlagwoerter)
            else:
                self.browser.open(
                    "https://www.handelsregister.de/rp_web/erweitertesuche.xhtml"
                )
                if self.args.debug == True:
                    print(self.browser.title())

                self.browser.select_form(name="form")

                self.browser["form:schlagwoerter"] = self.args.schlagwoerter
                so_id = schlagwortOptionen.get(self.args.schlagwortOptionen)

                self.browser["form:schlagwortOptionen"] = [str(so_id)]

                response_result = self.browser.submit()

                if self.args.debug == True:
                    print(self.browser.title())

                html = response_result.read().decode("utf-8")
                with open(cachename, "w") as f:
                    f.write(html) 

            return get_companies_in_searchresults(html) 

    

        # TODO get all documents attached to the exact company 


def get_companies_in_searchresults(html):
    soup = BeautifulSoup(html, "html.parser")
    grid = soup.find("table", role="grid")

    results = []
    for result in grid.find_all("tr"):
        a = result.get("data-ri")
        if a is not None:
            d = parse_result(result)
            results.append(d)
    return results 


# def pr_company_info(c):
#     # print(c)
#     for tag in ("name", "court", "state", "status"):
#         print("%s: %s" % (tag, c.get(tag, "-"))) 


def save_to_excel(firmenname, gericht, sitz, status, bezeichnung, rechtsform, strasse, hausnummer, postleitzahl, directorArray, gegenstand, vertretungsbefugnis, filepath): 
    # Überprüfen, ob die Datei existiert, ansonsten eine neue erstellen 
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active 

    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active 

        # Header einfügen
        sheet.append(
            [
                "Firmenname",
                "Gericht",
                "Sitz",
                "Status",
                "Bezeichnung", 
                "Rechtsform", 
                "Straße", 
                "Hausnummer", 
                "Postleitzahl", 
                "Ort", 
                "Geschäftsführer(in) Vorname", 
                "Geschäftsführer(in) Nachname", 
                "Geschäftsführer(in) Geschlecht", 
                "Geschäftsführer(in) Geburtsdatum", 
                "Gegenstand", 
                "Vertretungsbefugnis"
            ]
        ) 

    for director in directorArray: 
        sheet.append(
            [
                firmenname, 
                gericht, 
                sitz, 
                status, 
                bezeichnung, 
                rechtsform, 
                strasse, 
                hausnummer, 
                postleitzahl, 
                director[3], 
                director[0], 
                director[1], 
                " ",
                director[2], 
                gegenstand, 
                vertretungsbefugnis,
            ]
        )

    workbook.save(filepath)



def parse_result(result):
    cells = []
    for cellnum, cell in enumerate(result.find_all("td")):
        cells.append(cell.text.strip())

    d = {}
    d["court"] = cells[1]  # Gericht
    d["name"] = cells[2]  # Firmenname
    d["state"] = cells[3]  # Sitz
    d["status"] = cells[4]  # Status
    d["documents"] = cells[5]  # Dokumente

    # Extract and print the relevant portion of HTML containing "SI"
    if "SI" in cells[5]:
        # start_index = max(cells[5].find("SI") - 100, 0)
        # end_index = cells[5].find("SI") + 100
        # print("---- Debug: HTML Around 'SI' ----")
        # print(cells[5][start_index:end_index])
        # print("---------------------------------") 

        # 
        obtain_and_parse_detailed_results(cells[2], cells[1], cells[3], cells[4])
        # 

    # d["history"] = [6]  # Verlauf

    # # Extract history if available
    # history_cells = result.find_all("td")[8:]
    # if history_cells:
    #     for i in range(0, len(history_cells), 2):
    #         event = history_cells[i].text.strip()
    #         date = history_cells[i + 1].text.strip()
    #         d["history"].append((event, date))

    return d 


# Parse more useful information out of the XML file (specifically that from the "SI" link) 
def obtain_and_parse_detailed_results(firmenname, gericht, sitz, status):
    # current_directory = pathlib.Path.cwd() 
    current_directory = os.getcwd() 
    print("Current Directory:", current_directory) 

    options = webdriver.ChromeOptions()
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": current_directory,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        },
    ) 

    # Initialize Web Driver
    driver = webdriver.Chrome(options=options) 

    # Visit the Advanced Search page
    url = "https://www.handelsregister.de/rp_web/erweitertesuche.xhtml"
    driver.get(url) 

    # Fill out the form
    schlagwoerter = driver.find_element(By.NAME, "form:schlagwoerter")
    schlagwoerter.send_keys(firmenname) 

    #
    # schlagwortOptionen Section

    try:
        # Wait until the element is visible and clickable
        schlagwortOptionen_exact = WebDriverWait(driver, 10).until( 
            EC.element_to_be_clickable(
                (By.XPATH, '//label[@for="form:schlagwortOptionen:2"]')
            ) 
        )

        # Scroll the "schlagwortOptionen_exact" button into view
        driver.execute_script("arguments[0].scrollIntoView();", schlagwortOptionen_exact)

        # Move the mouse over the element
        actions = ActionChains(driver)
        actions.move_to_element(schlagwortOptionen_exact).perform()

        # Click on it
        schlagwortOptionen_exact.click() 

    except selenium.common.exceptions.ElementNotInteractableException:
        print("Element not interactable, trying JavaScript click.")
        driver.execute_script("arguments[0].click();", schlagwortOptionen_exact)

    except selenium.common.exceptions.TimeoutException:
        print("Element not found or not clickable within the wait time.") 


    #
    # Suchen Button Section

    try:

        # Locate the "Suchen" button and submit the form
        find_button = driver.find_element(By.NAME, "form:btnSuche")

        # Scroll the "Suchen" button into view
        driver.execute_script("arguments[0].scrollIntoView();", find_button)

        # Move the mouse over the element
        actions = ActionChains(driver)
        actions.move_to_element(find_button).perform()

        find_button.click()

    except selenium.common.exceptions.ElementClickInterceptedException:
        print("Click intercepted, trying JavaScript click.")
        driver.execute_script("arguments[0].click();", find_button)

    except selenium.common.exceptions.TimeoutException:
        print("Element not found or not clickable within the wait time.")



    # Wait again for the next page to load
    time.sleep(10)

    # Download the "SI" xhtml file
    download_button = driver.find_element(By.LINK_TEXT, "SI")
    download_button.click()

    # Wait for the download to complete
    time.sleep(10) 


    # Verify download
    # Get the latest downloaded file name
    files = os.listdir(current_directory)
    files = [f for f in files if not f.endswith(".crdownload")]  # Exclude incomplete files
    files.sort(key=lambda x: os.path.getctime(os.path.join(current_directory, x)))
    latest_file = files[-1]

    print(f"Downloaded file: {latest_file}")

    if os.name == "nt":
        print(
            f"Downloaded file directory (Windows notation): {current_directory + '\\' + latest_file}"
        )
    elif os.name == "posix":
        print(
            f"Downloaded file directory (Linux-like notation): {current_directory + '/' + latest_file}"
        )
    else:
        print(f"Downloaded file directory: {current_directory + '/' + latest_file}")


    # Close the browser
    driver.quit() 


    #
    #
    # Traversing the latest downloaded XML file and extracting data
    #
    #

    tree = ET.parse(current_directory + "\\" + latest_file)
    root = tree.getroot()

    namespaces = {"tns": "http://www.xjustiz.de"}  


    # Company Detailed Info 

    # Bezeichnung 
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:bezeichnung/tns:bezeichnung.aktuell",
        namespaces,
    ) is not None: 
        bezeichnung = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:bezeichnung/tns:bezeichnung.aktuell",
            namespaces,
        ).text
        # print(f"Bezeichnung: {bezeichnung}") 
    else: 
        bezeichnung = ""
        # print(f"Bezeichnung: {bezeichnung}") 

    # Rechtsform 
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:angabenZurRechtsform/tns:rechtsform/code",
        namespaces,
    ) is not None:
        rechtsform = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:angabenZurRechtsform/tns:rechtsform/code",
            namespaces,
        ).text
        if rechtsform == "GmbH": 
            rechtsform = "Gesellschaft mit beschränkter Haftung"
            # print(f"Rechtsform: \"Gesellschaft mit beschränkter Haftung\"")
        else: 
            rechtsform = ""
            # print(f"Rechtsform: \"Unbekannt\"") 
    else: 
        rechtsform = "" 
        print(f"Rechtsform: {rechtsform}") 
    
    # ort = root.find(
    #     "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:sitz/tens:ort",
    #     namespaces,
    # ).text
    # print(f"Ort: {ort}")
    
    # Strasse 
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:strasse",
        namespaces,
    ) is not None: 
        strasse = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:strasse",
            namespaces,
        ).text
        # print(f"Strasse: {strasse}") 
    else: 
        strasse = ""
        # print(f"Strasse: {strasse}")

    # Hausnummer 
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:hausnummer",
        namespaces,
    ) is not None: 
        hausnummer = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:hausnummer",
            namespaces,
        ).text
        # print(f"Hausnummer: {hausnummer}") 
    else: 
        hausnummer = ""
        # print(f"Hausnummer: {hausnummer}")

    # Postleitzahl
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:postleitzahl",
        namespaces,
    ) is not None:
        postleitzahl = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:postleitzahl",
            namespaces,
        ).text
        # print(f"Postleitzahl: {postleitzahl}") 
    else:
        postleitzahl = ""
        # print(f"Postleitzahl: {postleitzahl}")

    # ort = root.find(
    #     "tns:fachdatenRegister/tns:basisdatenRegister/tns:rechtstraeger/tns:anschrift/tns:ort",
    #     namespaces,
    # ).text
    # print(f"Ort: {ort}")

    # Gegenstand
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:gegenstand",
        namespaces,
    ) is not None:
        gegenstand = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:gegenstand",
            namespaces,
        ).text
        # print(f"Gegenstand: {gegenstand}") 
    else: 
        gegenstand = ""
        # print(f"Gegenstand: {gegenstand}") 

    # Vertretungsbefugnis
    if root.find(
        "tns:fachdatenRegister/tns:basisdatenRegister/tns:vertretung/tns:allgemeineVertretungsregelung/tns:auswahl_vertretungsbefugnis/tns:vertretungsbefugnisFreitext",
        namespaces,
    ) is not None: 
        vertretungsbefugnis = root.find(
            "tns:fachdatenRegister/tns:basisdatenRegister/tns:vertretung/tns:allgemeineVertretungsregelung/tns:auswahl_vertretungsbefugnis/tns:vertretungsbefugnisFreitext",
            namespaces,
        ).text 
        # print(f"Vertretungsbefugnis: {vertretungsbefugnis}")
    else: 
        vertretungsbefugnis = "" 
        # print(f"Vertretungsbefugnis: {vertretungsbefugnis}") 

    # # Iterate over all 'company' elements
    # for company in root.findall('tns:grunddaten/tns:verfahrensdaten/tns:beteiligung/tns:beteiligter/tns:auswahl_beteiligter/tns:organisation', namespaces,):
    #     company_name = company.find('tns:bezeichnung/tns:bezeichnung.aktuell', namespaces,).text 
    #     print(company_name)
    #     ort = company.find('tns:sitz/tns:ort', namespaces,).text
    #     print(ort) 


    # print("-----------------------------") 

    directorArray = []

    # Iterate over all 'director' elements
    for director in root.findall('tns:grunddaten/tns:verfahrensdaten/tns:beteiligung/tns:beteiligter/tns:auswahl_beteiligter/tns:natuerlichePerson', namespaces,): 

        # empty the variables first, if the loop has been run already
        vorname = ''
        nachname = '' 
        geburtsdatum = '' 
        ort = ''

        vorname = director.find('tns:vollerName/tns:vorname', namespaces,).text 
        # print(vorname)
        nachname = director.find('tns:vollerName/tns:nachname', namespaces,).text
        # print(nachname)
        geburtsdatum = director.find('tns:geburt/tns:geburtsdatum', namespaces,).text
        # print(geburtsdatum)
        ort = director.find('tns:anschrift/tns:ort', namespaces,).text
        # print(ort) 


        updatedDirectorArray = [ vorname, nachname, geburtsdatum, ort ]

        directorArray.append(updatedDirectorArray) 


        # print("----------------------------")
        # print(firmenname, gericht, sitz, status, bezeichnung, rechtsform, strasse, hausnummer, postleitzahl, ort, vorname, nachname, geburtsdatum, gegenstand, vertretungsbefugnis) 
        # print("------------------------------------------------") 

    # print(directorArray) 
    save_to_excel(firmenname, gericht, sitz, status, bezeichnung, rechtsform, strasse, hausnummer, postleitzahl, directorArray, gegenstand, vertretungsbefugnis, args.output) 



def parse_args():
    parser = argparse.ArgumentParser(description="A handelsregister CLI")
    parser.add_argument(
        "-d",
        "--debug",
        help="Enable debug mode and activate logging",
        action="store_true",
    )
    parser.add_argument(
        "-f",
        "--force",
        help="Force a fresh pull and skip the cache",
        action="store_true",
    )
    parser.add_argument(
        "-s",
        "--schlagwoerter",
        help="Search for the provided keywords",
        default="Kloepfel Consulting GmbH",
    )
    parser.add_argument(
        "-so",
        "--schlagwortOptionen",
        help="Keyword options: all=contain all keywords; min=contain at least one keyword; exact=contain the exact company name.",
        choices=["all", "min", "exact"],
        default="exact",
    )
    parser.add_argument(
        "-adv",
        "--advanced",
        help="Keyword options: true=retrieve more details about the company.; false=do not retrieve more details about the company.",
        choices=["true", "false"],
        default="true",
    )
    parser.add_argument(
        "-lxxlsx",
        "--loadExternalExcel",
        help="Keyword options (case-sensitive): true=loads an external excel file containing names of companies.; false=do not load an external excel file containing names of companies.",
        choices=["true", "false"],
        default="false",
    )
    parser.add_argument(
        "-xxlsx",
        "--externalExcel",
        help="Path to the external Excel file containing names of companies prepared for search. You can also specify the filename; and the filename must have a suffix of \".xlsx\". It must be contained within the \"external_excel\" folder.",
        default="handelsregister_query.xlsx",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Path to the output Excel file",
        default="handelsregister_result.xlsx",
    )
    args = parser.parse_args()

    if args.debug:
        import logging

        logger = logging.getLogger("mechanize")
        logger.addHandler(logging.StreamHandler(sys.stdout))
        logger.setLevel(logging.DEBUG)

    return args


if __name__ == "__main__":
    args = parse_args()
    h = HandelsRegister(args)
    h.open_startpage()
    companies = h.search_company_ies() 

    if companies is not None:
        # for c in companies:
            # pr_company_info(c) 
            # save_to_excel(companies, args.output)
        print(f"Ergebnisse wurden in der Datei {args.output} gespeichert.") 
